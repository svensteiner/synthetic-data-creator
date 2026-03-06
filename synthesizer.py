"""
synthesizer.py - Kern-Logik fuer synthetische Datengenerierung
Kein pandas/numpy - nur openpyxl + Faker fuer minimale EXE-Groesse.
Laeuft 100% lokal, keine Daten verlassen den Rechner.
"""

import datetime as dt
import logging
import re
import random
import math
from collections import defaultdict
from typing import Any, Callable

log = logging.getLogger(__name__)

from faker import Faker
from openpyxl import load_workbook

fake = Faker("de_DE")
Faker.seed(42)
random.seed(42)

_value_maps: dict[str, dict] = defaultdict(dict)


def reset_maps():
    _value_maps.clear()


def _consistent(col_key: str, original: Any, generator: Callable) -> Any:
    key = str(original).strip().lower()
    if key not in _value_maps[col_key]:
        _value_maps[col_key][key] = generator()
    return _value_maps[col_key][key]


# ---------------------------------------------------------------------------
# Spaltentyp-Erkennung (inkl. SAP-Kuerzel)
# ---------------------------------------------------------------------------

COLUMN_PATTERNS = {
    "vorname": [
        r"vorname", r"first.?name", r"v\.?name",
    ],
    "nachname": [
        r"nachname", r"last.?name", r"familienname", r"n\.?name",
    ],
    "vollname": [
        r"vollst.*name", r"full.?name", r"kundenname", r"^name$",
    ],
    "firma": [
        r"firma", r"unternehmen", r"gesellschaft", r"company", r"arbeitgeber",
        r"lieferant(?!.?nr)", r"mandant(?!.?nr)",
    ],
    "strasse": [
        r"stra[sß]e", r"adresse", r"address", r"anschrift", r"^str\.",
    ],
    "plz": [
        r"plz", r"postleitzahl", r"postal", r"zip",
    ],
    "ort": [
        r"^ort$", r"stadt", r"city", r"wohnort", r"gemeinde",
    ],
    "land": [
        r"^land$", r"country", r"staat",
    ],
    "email": [
        r"e.?mail", r"mail",
    ],
    "telefon": [
        r"telefon", r"^tel", r"handy", r"mobil", r"phone", r"fax", r"fon",
    ],
    "iban": [
        r"^iban$", r"^bank.?konto",
    ],
    "bic": [
        r"^bic$", r"swift",
    ],
    "steuernummer": [
        r"steuernummer", r"steuer.?nr", r"steuer.?num",
    ],
    "ustid": [
        r"ust.?id", r"umsatzsteuer.?id", r"vat.?id", r"mwst.?nr",
    ],
    "handelsreg": [
        r"handelsreg", r"^hrb", r"^hra", r"amtsgericht",
    ],
    "kundennummer": [
        r"kunden.?nr", r"kunden.?id", r"kundennummer", r"kd.?nr", r"customer.?id",
        r"deb.?nr", r"kred.?nr",
        r"lfd.?nr", r"laufende.?nr",
        r"liefnr", r"lieferanten.?nr",
        r"gesch.?partner", r"gp.?nr",
    ],
    "rechnungsnr": [
        r"rechnungs.?nr", r"rechnungsnummer", r"invoice.?nr",
        r"beleg.?nr", r"bel.?nr", r"buch.?nr",
        r"^re.?nr", r"^rg.?nr",
    ],
    "betrag": [
        r"betrag", r"summe", r"gesamt", r"^wert$", r"preis",
        r"netto", r"brutto", r"eur", r"amount",
        r"^btr", r"zahlbetrag", r"rechnungsbetrag",
        r"mwst.?betr", r"steuer.?betr",
        r"restbetrag", r"offener.?posten",
    ],
    "datum": [
        r"datum", r"date", r"zeitpunkt",
        r"buchungs", r"rechnungs.*dat", r"geburts",
        r"buch.?dat", r"val.?dat", r"wert.?dat",
        r"faell", r"faellig",
        r"lief.?dat", r"eingang",
    ],
    "beschreibung": [
        r"beschreibung", r"bezeichnung", r"betreff", r"kommentar", r"notiz", r"memo",
    ],
}

IBAN_REGEX  = re.compile(r"^[A-Z]{2}\d{2}[\dA-Z]{4,}$")
EMAIL_REGEX = re.compile(r"^[\w.+\-]+@[\w\-]+\.\w{2,}$")

ALL_TYPES = [
    "text",
    "vorname", "nachname", "vollname",
    "firma",
    "strasse", "plz", "ort", "land",
    "email", "telefon",
    "iban", "bic",
    "steuernummer", "ustid", "handelsreg",
    "kundennummer", "rechnungsnr",
    "betrag", "datum",
    "beschreibung",
]

TYPE_LABELS = {
    "text":         "--- Nicht anonymisieren ---",
    "vorname":      "Vorname",
    "nachname":     "Nachname",
    "vollname":     "Vollstaendiger Name",
    "firma":        "Firma / Unternehmen",
    "strasse":      "Strasse / Adresse",
    "plz":          "PLZ",
    "ort":          "Ort / Stadt",
    "land":         "Land",
    "email":        "E-Mail",
    "telefon":      "Telefon / Handy",
    "iban":         "IBAN",
    "bic":          "BIC / SWIFT",
    "steuernummer": "Steuernummer",
    "ustid":        "USt-IdNr",
    "handelsreg":   "Handelsregisternummer",
    "kundennummer": "Kundennummer / ID",
    "rechnungsnr":  "Rechnungsnummer",
    "betrag":       "Betrag / Zahl",
    "datum":        "Datum",
    "beschreibung": "Beschreibung (unveraendert)",
}

LABEL_TO_TYPE = {v: k for k, v in TYPE_LABELS.items()}


def _is_nan(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def detect_column_type(col_name: str, values: list) -> str:
    col_lower = col_name.lower().strip()

    for col_type, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, col_lower):
                return col_type

    sample = [str(v) for v in values if not _is_nan(v)][:30]

    if sample:
        if sum(1 for v in sample if IBAN_REGEX.match(v)) / len(sample) > 0.5:
            return "iban"
        if sum(1 for v in sample if EMAIL_REGEX.match(v)) / len(sample) > 0.5:
            return "email"

    non_null = [v for v in values if not _is_nan(v)]
    if non_null:
        if all(isinstance(v, (dt.datetime, dt.date)) for v in non_null):
            return "datum"
        if all(isinstance(v, (int, float)) for v in non_null):
            return "betrag"

    return "text"


# ---------------------------------------------------------------------------
# Analyse (ohne Generierung) - fuer Vorschau-Dialog
# ---------------------------------------------------------------------------

def _find_header_row(ws) -> int:
    """
    Findet die echte Header-Zeile automatisch.
    Ueberspringt Titelzeilen (z.B. 'BILANZ ZUM 31.12.2024' in Zeile 1).
    Kriterium: erste Zeile mit mind. 2 String-Werten in >= 40% der Spalten.
    """
    n_cols = max(ws.max_column or 1, 1)
    for row in range(1, min(10, (ws.max_row or 1) + 1)):
        values  = [ws.cell(row, c).value for c in range(1, n_cols + 1)]
        strings = [v for v in values if isinstance(v, str) and v.strip()]
        if len(strings) >= 2 and len(strings) / n_cols >= 0.4:
            return row
    return 1


def analyze_excel(input_path: str) -> dict:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws      = wb[sheet_name]
        rows    = list(ws.iter_rows(values_only=True))
        wb_info = {}

        if not rows:
            result[sheet_name] = wb_info
            continue

        hdr_idx   = _find_header_row(ws) - 1          # 0-based index into rows
        headers   = [str(h).strip() if h is not None else f"Spalte{i}"
                     for i, h in enumerate(rows[hdr_idx], 1)]
        data_rows = rows[hdr_idx + 1:]

        for col_idx, col_name in enumerate(headers):
            values   = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
            col_type = detect_column_type(col_name, values)
            samples  = [str(v) for v in values if not _is_nan(v)][:3]
            wb_info[col_name] = {"type": col_type, "samples": samples}

        result[sheet_name] = wb_info

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Fake-Wert-Generatoren
# ---------------------------------------------------------------------------

def _fake_german_iban() -> str:
    bban = "".join(str(random.randint(0, 9)) for _ in range(18))
    return f"DE{random.randint(10, 99)}{bban}"


def _fake_steuernummer() -> str:
    return f"{random.randint(10,99)}/{random.randint(100,999)}/{random.randint(10000,99999)}"


def _fake_ustid() -> str:
    return f"DE{random.randint(100_000_000, 999_999_999)}"


def _fake_handelsreg() -> str:
    return f"{random.choice(['HRB', 'HRA'])} {random.randint(1000, 999999)}"


def _fake_kundennummer(original: str) -> str:
    digits = re.sub(r"\D", "", original)
    length = max(len(digits), 5)
    return str(random.randint(10 ** (length - 1), 10**length - 1))


def _fake_rechnungsnr(original: str) -> str:
    m      = re.match(r"^([A-Za-z\-/]+\d{2,4}[\-/]?)", original)
    prefix = m.group(1) if m else "RE-"
    return f"{prefix}{random.randint(1000, 99999)}"


def _fake_betrag(original: Any, stats: dict) -> Any:
    try:
        val = float(original)
    except (ValueError, TypeError):
        return original

    min_v    = stats.get("min", val)
    max_v    = stats.get("max", val)
    decimals = stats.get("decimals", 2)

    if abs(max_v - min_v) < 0.01:
        return round(val * random.uniform(0.88, 1.12), decimals)

    return round(random.uniform(min_v, max_v), decimals)


def _fake_datum(original: Any) -> Any:
    delta = random.randint(-180, 180)
    if isinstance(original, (dt.date, dt.datetime)):
        return original + dt.timedelta(days=delta)
    return original


def generate_fake_value(col_type: str, original: Any, col_key: str, stats: dict | None = None) -> Any:
    if _is_nan(original):
        return original

    stats = stats or {}

    match col_type:
        case "vorname":
            return _consistent(col_key, original, fake.first_name)
        case "nachname":
            return _consistent(col_key, original, fake.last_name)
        case "vollname":
            return _consistent(col_key, original, lambda: f"{fake.first_name()} {fake.last_name()}")
        case "firma":
            return _consistent(col_key, original, fake.company)
        case "strasse":
            return _consistent(col_key, original, lambda: f"{fake.street_name()} {random.randint(1, 150)}")
        case "plz":
            return _consistent(col_key, original, fake.postcode)
        case "ort":
            return _consistent(col_key, original, fake.city)
        case "land":
            return _consistent(col_key, original, fake.country)
        case "email":
            return _consistent(col_key, original, fake.email)
        case "telefon":
            return _consistent(col_key, original, fake.phone_number)
        case "iban":
            return _consistent(col_key, original, _fake_german_iban)
        case "bic":
            return _consistent(col_key, original, lambda: fake.swift(length=8))
        case "steuernummer":
            return _consistent(col_key, original, _fake_steuernummer)
        case "ustid":
            return _consistent(col_key, original, _fake_ustid)
        case "handelsreg":
            return _consistent(col_key, original, _fake_handelsreg)
        case "kundennummer":
            return _consistent(col_key, original, lambda: _fake_kundennummer(str(original)))
        case "rechnungsnr":
            return _consistent(col_key, original, lambda: _fake_rechnungsnr(str(original)))
        case "betrag":
            return _fake_betrag(original, stats)
        case "datum":
            return _fake_datum(original)
        case _:
            return original


# ---------------------------------------------------------------------------
# Statistiken fuer Betrag-Spalten (pure Python, kein numpy)
# ---------------------------------------------------------------------------

def _calc_stats(values: list) -> dict:
    nums     = []
    decimals = 0
    for v in values:
        if _is_nan(v):
            continue
        try:
            f = float(v)
            nums.append(f)
            s = f"{f:.10f}".rstrip("0")
            if "." in s:
                decimals = max(decimals, len(s.split(".")[-1]))
        except (ValueError, TypeError):
            pass

    if not nums:
        return {}

    return {
        "min":      min(nums),
        "max":      max(nums),
        "decimals": min(decimals, 4),
    }


# ---------------------------------------------------------------------------
# Hauptfunktion - openpyxl direkt (Formatierung bleibt erhalten!)
# ---------------------------------------------------------------------------

def synthesize_excel(
    input_path: str,
    output_path: str,
    overrides: dict[str, str] | None = None,
    progress_callback=None,
) -> dict:
    reset_maps()
    overrides = overrides or {}

    wb    = load_workbook(input_path)
    info  = {"sheets": {}, "row_counts": {}}
    total = len(wb.sheetnames)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        if progress_callback:
            progress_callback(sheet_idx / total, f"Verarbeite: {sheet_name}")

        ws        = wb[sheet_name]
        data_rows = ws.max_row - 1
        info["row_counts"][sheet_name] = max(data_rows, 0)

        if ws.max_row < 2 or ws.max_column < 1:
            info["sheets"][sheet_name] = {}
            continue

        # Header-Zeile automatisch finden (ueberspringt Titelzeilen)
        hdr_row = _find_header_row(ws)

        headers: dict[int, str] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=hdr_row, column=col_idx).value
            if val is not None:
                headers[col_idx] = str(val).strip()

        # Spaltenwerte als Listen einlesen (nur Datenzeilen)
        col_values: dict[int, list] = {}
        for col_idx in headers:
            col_values[col_idx] = [
                ws.cell(row=r, column=col_idx).value
                for r in range(hdr_row + 1, ws.max_row + 1)
            ]

        # Typen + Stats ermitteln
        col_types: dict[int, str] = {}
        col_stats: dict[int, dict] = {}
        sheet_info: dict[str, str] = {}

        for col_idx, col_name in headers.items():
            values   = col_values[col_idx]
            col_type = overrides.get(col_name) or detect_column_type(col_name, values)
            col_types[col_idx]   = col_type
            sheet_info[col_name] = col_type
            if col_type == "betrag":
                col_stats[col_idx] = _calc_stats(values)

        info["sheets"][sheet_name] = sheet_info

        # Werte in-place ersetzen (Formatierung bleibt!)
        for col_idx, col_name in headers.items():
            col_type = col_types[col_idx]
            if col_type in ("text", "beschreibung"):
                continue

            stats = col_stats.get(col_idx, {})

            for row_idx in range(hdr_row + 1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                try:
                    cell.value = generate_fake_value(col_type, cell.value, col_name, stats)
                except Exception as exc:
                    log.warning("Zelle (%s, %s) Zeile %d konnte nicht anonymisiert werden: %s",
                                sheet_name, col_name, row_idx, exc)

    wb.save(output_path)

    if progress_callback:
        progress_callback(1.0, "Fertig!")

    return info
