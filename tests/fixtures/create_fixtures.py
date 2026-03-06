"""
create_fixtures.py
Erstellt realistische Test-Excel-Dateien (Bilanz, GuV, Offene Posten)
mit typischer Wirtschaftspruefer-Formatierung.
"""

import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Farben
C_HEADER    = "1A3A5C"
C_SUBHEAD   = "D6E4F0"
C_ROW_ALT   = "F5F7FA"
C_ROW_TOTAL = "E8EFF7"
C_GREEN_ALT = "F0FFF0"
C_WHITE     = "FFFFFF"
C_WARN      = "FFF3CD"

EURO_FMT  = '#,##0.00 €'
DATE_FMT  = 'DD.MM.YYYY'

def _side(style="thin"):
    return Side(style=style, color="CCCCCC")

def _border(style="thin"):
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)

def _header(ws, row, n_cols, bg=C_HEADER, fg="FFFFFF", height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(bold=True, color=fg, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()

def _data(ws, row, n_cols, bg=C_WHITE):
    ws.row_dimensions[row].height = 16
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center")
        cell.border    = _border()

def _total(ws, row, n_cols):
    ws.row_dimensions[row].height = 18
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = Font(bold=True, name="Calibri", size=10, color=C_HEADER)
        cell.fill   = PatternFill("solid", fgColor=C_ROW_TOTAL)
        cell.border = Border(
            top=Side(style="medium", color=C_HEADER),
            bottom=Side(style="medium", color=C_HEADER),
            left=_side(), right=_side()
        )

def _section(ws, row, label, n_cols):
    """Abschnittszeile z.B. AKTIVA / PASSIVA."""
    ws.row_dimensions[row].height = 18
    ws.cell(row=row, column=1, value=label)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
        cell.fill   = PatternFill("solid", fgColor="2C5F8A")
        cell.border = _border()


# ---------------------------------------------------------------------------
# Bilanz
# ---------------------------------------------------------------------------

def create_bilanz(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilanz"

    # Titelzeile
    ws.merge_cells("A1:E1")
    ws.cell(1, 1, "BILANZ ZUM 31. DEZEMBER 2024").font = Font(
        bold=True, size=14, color="FFFFFF", name="Calibri"
    )
    ws.cell(1, 1).fill      = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Unterzeile Mandant
    ws.merge_cells("A2:E2")
    ws.cell(2, 1, "Mustermann Holding GmbH & Co. KG  |  Steuer-Nr: 21/815/08150  |  HRB 123456")
    ws.cell(2, 1).font      = Font(italic=True, size=9, color="FFFFFF", name="Calibri")
    ws.cell(2, 1).fill      = PatternFill("solid", fgColor="2C5F8A")
    ws.cell(2, 1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 15

    # Spaltenheader
    for i, h in enumerate(["Konto-Nr", "Kontobezeichnung", "2024 (EUR)", "2023 (EUR)", "Veraenderung"], 1):
        ws.cell(3, i, h)
    _header(ws, 3, 5)

    # Aktiva
    _section(ws, 4, "A K T I V A", 5)

    aktiva = [
        ("0200", "Sachanlagen",                          125_000.00, 118_000.00),
        ("0500", "Finanzanlagen",                         45_000.00,  45_000.00),
        ("1000", "Kasse",                                 15_234.00,  12_100.00),
        ("1200", "Bankguthaben - Commerzbank",           248_500.00, 195_700.00),
        ("1201", "Bankguthaben - Sparkasse",              32_100.00,  28_000.00),
        ("1400", "Forderungen aus Lieferungen & Leist.",  87_320.00,  95_400.00),
        ("1600", "Vorraete",                              32_100.00,  28_800.00),
        ("1800", "Sonstige Forderungen",                   9_200.00,   7_500.00),
        ("1900", "Aktive Rechnungsabgrenzung",             3_800.00,   2_900.00),
    ]
    for i, (konto, bez, b24, b23) in enumerate(aktiva):
        r   = i + 5
        bg  = C_ROW_ALT if i % 2 == 0 else C_WHITE
        ws.cell(r, 1, konto)
        ws.cell(r, 2, bez)
        ws.cell(r, 3, b24)
        ws.cell(r, 4, b23)
        ws.cell(r, 5, round(b24 - b23, 2))
        _data(ws, r, 5, bg)
        for c in [3, 4, 5]:
            ws.cell(r, c).number_format = EURO_FMT
            ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")

    r_sum_a = len(aktiva) + 5
    ws.cell(r_sum_a, 2, "SUMME AKTIVA")
    sum_a24 = sum(x[2] for x in aktiva)
    sum_a23 = sum(x[3] for x in aktiva)
    ws.cell(r_sum_a, 3, sum_a24)
    ws.cell(r_sum_a, 4, sum_a23)
    ws.cell(r_sum_a, 5, round(sum_a24 - sum_a23, 2))
    _total(ws, r_sum_a, 5)
    for c in [3, 4, 5]:
        ws.cell(r_sum_a, c).number_format = EURO_FMT
        ws.cell(r_sum_a, c).alignment = Alignment(horizontal="right", vertical="center")

    # Passiva
    r_pass = r_sum_a + 1
    _section(ws, r_pass, "P A S S I V A", 5)

    passiva = [
        ("0800", "Eigenkapital",                         180_000.00, 160_000.00),
        ("0850", "Kapitalruecklage",                      45_000.00,  40_000.00),
        ("0900", "Gewinnruecklagen",                      22_500.00,  20_000.00),
        ("0990", "Jahresueberschuss",                     25_354.00,  22_100.00),
        ("3200", "Rueckstellungen",                       18_200.00,  16_500.00),
        ("3300", "Verbindlichkeiten aus LuL",             98_540.00, 102_300.00),
        ("3500", "Verbindlichkeiten gegenueber Banken",   75_000.00,  80_000.00),
        ("3700", "Sonstige Verbindlichkeiten",            28_760.00,  25_400.00),
        ("3900", "Passive Rechnungsabgrenzung",            4_900.00,   3_100.00),
    ]
    for i, (konto, bez, b24, b23) in enumerate(passiva):
        r   = r_pass + 1 + i
        bg  = C_ROW_ALT if i % 2 == 0 else C_WHITE
        ws.cell(r, 1, konto)
        ws.cell(r, 2, bez)
        ws.cell(r, 3, b24)
        ws.cell(r, 4, b23)
        ws.cell(r, 5, round(b24 - b23, 2))
        _data(ws, r, 5, bg)
        for c in [3, 4, 5]:
            ws.cell(r, c).number_format = EURO_FMT
            ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")

    r_sum_p = r_pass + len(passiva) + 1
    ws.cell(r_sum_p, 2, "SUMME PASSIVA")
    sum_p24 = sum(x[2] for x in passiva)
    sum_p23 = sum(x[3] for x in passiva)
    ws.cell(r_sum_p, 3, sum_p24)
    ws.cell(r_sum_p, 4, sum_p23)
    ws.cell(r_sum_p, 5, round(sum_p24 - sum_p23, 2))
    _total(ws, r_sum_p, 5)
    for c in [3, 4, 5]:
        ws.cell(r_sum_p, c).number_format = EURO_FMT
        ws.cell(r_sum_p, c).alignment = Alignment(horizontal="right", vertical="center")

    # Spaltenbreiten
    for col, w in [("A", 12), ("B", 40), ("C", 18), ("D", 18), ("E", 16)]:
        ws.column_dimensions[col].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# GuV
# ---------------------------------------------------------------------------

def create_guv(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "GuV"

    # Titel
    ws.merge_cells("A1:D1")
    ws.cell(1, 1, "GEWINN- UND VERLUSTRECHNUNG  |  01.01.2024 – 31.12.2024")
    ws.cell(1, 1).font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    ws.cell(1, 1).fill      = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws.cell(2, 1, "Mustermann Holding GmbH & Co. KG  |  nach § 275 HGB (Gesamtkostenverfahren)")
    ws.cell(2, 1).font      = Font(italic=True, size=9, color="FFFFFF", name="Calibri")
    ws.cell(2, 1).fill      = PatternFill("solid", fgColor="2C5F8A")
    ws.cell(2, 1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    for i, h in enumerate(["Pos.", "Bezeichnung", "2024 (EUR)", "2023 (EUR)"], 1):
        ws.cell(3, i, h)
    _header(ws, 3, 4)

    guv_rows = [
        # (pos, bezeichnung, 2024, 2023, is_total)
        ("1.",  "Umsatzerloese",                                    1_250_000.00, 1_100_000.00, False),
        ("2.",  "Erhoehung Bestand fertiger Erzeugnisse",               15_000.00,    12_000.00, False),
        ("3.",  "Sonstige betriebliche Ertraege",                        28_500.00,    24_000.00, False),
        (None,  "BETRIEBSERTRAG",                                    1_293_500.00, 1_136_000.00, True),
        ("4.",  "Materialaufwand",                                    -450_000.00,  -398_000.00, False),
        ("5.",  "Personalaufwand",                                    -380_000.00,  -355_000.00, False),
        ("6.",  "Abschreibungen auf Sachanlagen",                      -45_000.00,   -42_000.00, False),
        ("7.",  "Sonstige betriebliche Aufwendungen",                   -98_000.00,   -89_500.00, False),
        (None,  "BETRIEBSERGEBNIS (EBIT)",                             320_500.00,   251_500.00, True),
        ("8.",  "Zinsertraege",                                          3_200.00,     2_800.00, False),
        ("9.",  "Zinsaufwand",                                         -12_500.00,   -11_000.00, False),
        (None,  "ERGEBNIS VOR STEUERN (EBT)",                          311_200.00,   243_300.00, True),
        ("10.", "Steuern vom Einkommen und vom Ertrag",                -83_000.00,   -65_000.00, False),
        ("11.", "Sonstige Steuern",                                     -2_846.00,    -2_700.00, False),
        (None,  "JAHRESUEBERSCHUSS",                                   225_354.00,   175_600.00, True),
    ]

    for i, (pos, bez, b24, b23, is_total) in enumerate(guv_rows):
        r = i + 4
        ws.cell(r, 1, pos or "")
        ws.cell(r, 2, bez)
        ws.cell(r, 3, b24)
        ws.cell(r, 4, b23)
        if is_total:
            _total(ws, r, 4)
        else:
            bg = C_ROW_ALT if i % 2 == 0 else C_WHITE
            _data(ws, r, 4, bg)
        for c in [3, 4]:
            ws.cell(r, c).number_format = EURO_FMT
            ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")

    for col, w in [("A", 6), ("B", 48), ("C", 18), ("D", 18)]:
        ws.column_dimensions[col].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Offene Posten
# ---------------------------------------------------------------------------

def create_offene_posten(path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Offene Posten"

    ws.merge_cells("A1:H1")
    ws.cell(1, 1, "OFFENE-POSTEN-LISTE  |  Stand: 31.12.2024")
    ws.cell(1, 1).font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws.cell(1, 1).fill      = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["KdNr", "Kundenname", "IBAN", "Re-Nr", "Rechnungsdatum",
               "Faelligkeitsdatum", "Betrag (EUR)", "Status"]
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    _header(ws, 2, 8)

    rows = [
        ("10001", "Mueller GmbH",         "DE89370400440532013000", "RE-2024-10045",
         datetime.date(2024, 10,  1), datetime.date(2024, 10, 31), 12_450.00, "offen"),
        ("10002", "Schmidt & Partner KG", "DE12500105170648489890", "RE-2024-10067",
         datetime.date(2024, 10, 15), datetime.date(2024, 11, 14),  8_200.00, "offen"),
        ("10001", "Mueller GmbH",         "DE89370400440532013000", "RE-2024-10089",
         datetime.date(2024, 11,  1), datetime.date(2024, 12,  1),  5_750.00, "offen"),
        ("10003", "Wagner Holding AG",    "DE75512108001245126199", "RE-2024-10102",
         datetime.date(2024, 11, 20), datetime.date(2024, 12, 20), 34_800.00, "ueberfaellig"),
        ("10004", "Becker KG",            "DE09750100000012345678", "RE-2024-10115",
         datetime.date(2024, 12,  1), datetime.date(2024, 12, 31),  2_100.00, "offen"),
        ("10002", "Schmidt & Partner KG", "DE12500105170648489890", "RE-2024-10134",
         datetime.date(2024, 12, 10), datetime.date(2025,  1,  9), 15_300.00, "offen"),
        ("10005", "Hoffmann Logistik",    "DE21700100800030982700", "RE-2024-10156",
         datetime.date(2024, 12, 15), datetime.date(2025,  1, 14),  7_650.00, "offen"),
        ("10006", "Richter & Soehn GmbH", "DE81200400600100500700", "RE-2024-10178",
         datetime.date(2024, 12, 20), datetime.date(2025,  1, 19), 21_900.00, "offen"),
    ]

    for i, (kdnr, name, iban, renr, rdat, fdat, betrag, status) in enumerate(rows):
        r  = i + 3
        bg = C_WARN if status == "ueberfaellig" else (C_GREEN_ALT if i % 2 == 0 else C_WHITE)
        vals = [kdnr, name, iban, renr, rdat, fdat, betrag, status]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
        _data(ws, r, 8, bg)
        ws.cell(r, 7).number_format = EURO_FMT
        ws.cell(r, 7).alignment = Alignment(horizontal="right", vertical="center")
        for c in [5, 6]:
            ws.cell(r, c).number_format = DATE_FMT

    for col, w in zip("ABCDEFGH", [10, 24, 28, 18, 16, 18, 16, 12]):
        ws.column_dimensions[col].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Alle Fixtures erstellen
# ---------------------------------------------------------------------------

def create_all(base_dir: str | None = None) -> dict[str, str]:
    base = Path(base_dir) if base_dir else Path(__file__).parent
    paths = {
        "bilanz":        str(base / "bilanz.xlsx"),
        "guv":           str(base / "guv.xlsx"),
        "offene_posten": str(base / "offene_posten.xlsx"),
    }
    create_bilanz(paths["bilanz"])
    create_guv(paths["guv"])
    create_offene_posten(paths["offene_posten"])
    print(f"Fixtures erstellt in: {base}")
    for name, path in paths.items():
        size = Path(path).stat().st_size
        print(f"  {name:20s} {size:>8,} Bytes")
    return paths


if __name__ == "__main__":
    create_all()
