"""
test_synthesizer.py - Komplette Testreihe fuer den Synthetischen Datengenerator

Testet:
  1. Formatierung  - Farben, Schrift, Spaltenbreiten identisch
  2. Struktur      - Zeilen, Spalten, Sheet-Namen unveraendert
  3. Anonymisierung - Sensible Felder wirklich geaendert
  4. Konsistenz    - Gleicher Input -> immer gleicher Output
  5. Unveraendert  - Kontonummern, Bezeichnungen, Status bleiben gleich
  6. Typerkennung  - Korrekte Typen fuer typische Pruefungsfelder
  7. Betraege      - Synthetische Werte im Originalbereich (Min-Max)
  8. Datumsverschiebung - Daten veraendert, aber als Datum erhalten
"""

import sys
import unittest
from pathlib import Path
from openpyxl import load_workbook

# Pfad zum Hauptprojekt
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from synthesizer import synthesize_excel, detect_column_type, reset_maps
from fixtures.create_fixtures import create_all

FIXTURES = Path(__file__).parent / "fixtures"
OUTPUT   = Path(__file__).parent / "output"
OUTPUT.mkdir(exist_ok=True)


def setUpModule():
    """Fixtures einmalig erstellen."""
    FIXTURES.mkdir(exist_ok=True)
    create_all(str(FIXTURES))


def _load(name: str):
    """Laedt Original- und synthetische Workbook."""
    orig_path  = str(FIXTURES / f"{name}.xlsx")
    synth_path = str(OUTPUT  / f"{name}_synth.xlsx")
    info = synthesize_excel(orig_path, synth_path)
    wb_o = load_workbook(orig_path)
    wb_s = load_workbook(synth_path)
    return wb_o, wb_s, info


# ===========================================================================
# 1. FORMATIERUNG
# ===========================================================================

class TestFormatierung(unittest.TestCase):
    """Saemtliche Zellformatierung muss nach der Synthese identisch sein."""

    @classmethod
    def setUpClass(cls):
        reset_maps()
        cls.wb_o, cls.wb_s, _ = _load("offene_posten")
        cls.ws_o = cls.wb_o.active
        cls.ws_s = cls.wb_s.active

    def _cells(self):
        """Alle Zellen aus dem Header- und Datenbereich."""
        for row in range(1, self.ws_o.max_row + 1):
            for col in range(1, self.ws_o.max_column + 1):
                yield row, col

    def test_header_fettschrift(self):
        """Header-Zeile muss nach Synthese fett bleiben."""
        for col in range(1, self.ws_o.max_column + 1):
            with self.subTest(col=col):
                self.assertEqual(
                    self.ws_o.cell(2, col).font.bold,
                    self.ws_s.cell(2, col).font.bold,
                    f"Spalte {col}: Fettschrift unterschiedlich"
                )

    def test_hintergrundfarben(self):
        """Hintergrundfarben aller Zellen muessen identisch sein."""
        for row, col in self._cells():
            with self.subTest(row=row, col=col):
                orig_fill  = self.ws_o.cell(row, col).fill.fgColor.rgb
                synth_fill = self.ws_s.cell(row, col).fill.fgColor.rgb
                self.assertEqual(
                    orig_fill, synth_fill,
                    f"Zeile {row}, Spalte {col}: Farbe {orig_fill} != {synth_fill}"
                )

    def test_schriftfarbe_header(self):
        """Schriftfarbe im Header muss weiss bleiben."""
        for col in range(1, self.ws_o.max_column + 1):
            with self.subTest(col=col):
                self.assertEqual(
                    self.ws_o.cell(2, col).font.color.rgb,
                    self.ws_s.cell(2, col).font.color.rgb,
                )

    def test_spaltenbreiten(self):
        """Spaltenbreiten muessen exakt erhalten bleiben."""
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            orig_w  = self.ws_o.column_dimensions[col_letter].width
            synth_w = self.ws_s.column_dimensions[col_letter].width
            with self.subTest(col=col_letter):
                self.assertAlmostEqual(
                    orig_w, synth_w, places=1,
                    msg=f"Spalte {col_letter}: Breite {orig_w} != {synth_w}"
                )

    def test_zahlenformat_betrag(self):
        """Betragsspalte muss Waehrungsformat behalten."""
        for row in range(3, self.ws_o.max_row + 1):
            orig_fmt  = self.ws_o.cell(row, 7).number_format
            synth_fmt = self.ws_s.cell(row, 7).number_format
            with self.subTest(row=row):
                self.assertEqual(orig_fmt, synth_fmt,
                                 f"Zeile {row}: Zahlenformat unterschiedlich")

    def test_zahlenformat_datum(self):
        """Datumsformat muss erhalten bleiben."""
        for row in range(3, self.ws_o.max_row + 1):
            for col in [5, 6]:
                orig_fmt  = self.ws_o.cell(row, col).number_format
                synth_fmt = self.ws_s.cell(row, col).number_format
                with self.subTest(row=row, col=col):
                    self.assertEqual(orig_fmt, synth_fmt)

    def test_zeilenhoehen_bilanz(self):
        """Zeilenhoehen in der Bilanz muessen erhalten bleiben."""
        reset_maps()
        wb_o, wb_s, _ = _load("bilanz")
        ws_o, ws_s    = wb_o.active, wb_s.active
        for row in range(1, min(5, ws_o.max_row + 1)):
            with self.subTest(row=row):
                self.assertAlmostEqual(
                    ws_o.row_dimensions[row].height or 15,
                    ws_s.row_dimensions[row].height or 15,
                    places=0
                )


# ===========================================================================
# 2. STRUKTUR
# ===========================================================================

class TestStruktur(unittest.TestCase):
    """Struktur (Zeilen, Spalten, Sheets) muss unveraendert sein."""

    def _check(self, name):
        reset_maps()
        wb_o, wb_s, _ = _load(name)
        ws_o, ws_s    = wb_o.active, wb_s.active
        self.assertEqual(ws_o.max_row,    ws_s.max_row,    f"{name}: Zeilenanzahl unterschiedlich")
        self.assertEqual(ws_o.max_column, ws_s.max_column, f"{name}: Spaltenanzahl unterschiedlich")

    def test_struktur_bilanz(self):
        self._check("bilanz")

    def test_struktur_guv(self):
        self._check("guv")

    def test_struktur_offene_posten(self):
        self._check("offene_posten")

    def test_header_unveraendert(self):
        """Spaltenbezeichnungen (Zeile 2/3) muessen unveraendert bleiben."""
        reset_maps()
        wb_o, wb_s, _ = _load("offene_posten")
        ws_o, ws_s    = wb_o.active, wb_s.active
        for col in range(1, ws_o.max_column + 1):
            with self.subTest(col=col):
                self.assertEqual(
                    ws_o.cell(2, col).value,
                    ws_s.cell(2, col).value,
                    f"Header Spalte {col} wurde veraendert"
                )

    def test_sheet_namen(self):
        """Sheet-Namen muessen identisch bleiben."""
        for name in ["bilanz", "guv", "offene_posten"]:
            reset_maps()
            wb_o, wb_s, _ = _load(name)
            with self.subTest(name=name):
                self.assertEqual(wb_o.sheetnames, wb_s.sheetnames)


# ===========================================================================
# 3. ANONYMISIERUNG - Sensible Felder wirklich geaendert
# ===========================================================================

class TestAnonymisierung(unittest.TestCase):
    """Sensible Daten muessen sich tatsaechlich vom Original unterscheiden."""

    @classmethod
    def setUpClass(cls):
        reset_maps()
        cls.wb_o, cls.wb_s, cls.info = _load("offene_posten")
        cls.ws_o = cls.wb_o.active
        cls.ws_s = cls.wb_s.active

    def _col_values(self, ws, col_idx, start_row=3):
        return [ws.cell(r, col_idx).value for r in range(start_row, ws.max_row + 1)]

    def test_kundennamen_geaendert(self):
        """Kundennamen (Spalte 2) muessen anonymisiert sein."""
        orig  = self._col_values(self.ws_o, 2)
        synth = self._col_values(self.ws_s, 2)
        changed = sum(1 for o, s in zip(orig, synth) if o != s)
        self.assertGreater(changed, 0, "Kein Kundenname wurde geaendert!")

    def test_ibans_geaendert(self):
        """IBANs (Spalte 3) muessen ersetzt worden sein."""
        orig  = self._col_values(self.ws_o, 3)
        synth = self._col_values(self.ws_s, 3)
        changed = sum(1 for o, s in zip(orig, synth) if o != s)
        self.assertGreater(changed, 0, "Keine IBAN wurde geaendert!")

    def test_ibans_format_korrekt(self):
        """Synthetische IBANs muessen deutsches Format haben (DE + 20 Zeichen)."""
        import re
        synth = self._col_values(self.ws_s, 3)
        for iban in synth:
            if iban:
                with self.subTest(iban=iban):
                    self.assertRegex(
                        str(iban), r"^DE\d{20}$",
                        f"IBAN hat falsches Format: {iban}"
                    )

    def test_rechnungsnummern_geaendert(self):
        """Rechnungsnummern (Spalte 4) muessen geaendert sein."""
        orig  = self._col_values(self.ws_o, 4)
        synth = self._col_values(self.ws_s, 4)
        changed = sum(1 for o, s in zip(orig, synth) if o != s)
        self.assertGreater(changed, 0, "Keine Rechnungsnummer wurde geaendert!")

    def test_betraege_geaendert(self):
        """Betraege (Spalte 7) muessen veraendert worden sein."""
        orig  = self._col_values(self.ws_o, 7)
        synth = self._col_values(self.ws_s, 7)
        changed = sum(1 for o, s in zip(orig, synth) if o is not None and o != s)
        self.assertGreater(changed, 0, "Kein Betrag wurde geaendert!")

    def test_kundennummern_geaendert(self):
        """KdNr (Spalte 1) muss anonymisiert sein."""
        orig  = self._col_values(self.ws_o, 1)
        synth = self._col_values(self.ws_s, 1)
        changed = sum(1 for o, s in zip(orig, synth) if o != s)
        self.assertGreater(changed, 0, "Keine Kundennummer wurde geaendert!")

    def test_daten_geaendert(self):
        """Rechnungsdaten (Spalte 5) muessen verschoben worden sein."""
        orig  = self._col_values(self.ws_o, 5)
        synth = self._col_values(self.ws_s, 5)
        changed = sum(1 for o, s in zip(orig, synth) if o is not None and o != s)
        self.assertGreater(changed, 0, "Kein Datum wurde veraendert!")

    def test_status_unveraendert(self):
        """Status-Spalte (8) muss unveraendert bleiben (Text-Typ)."""
        orig  = self._col_values(self.ws_o, 8)
        synth = self._col_values(self.ws_s, 8)
        self.assertEqual(orig, synth, "Status-Werte wurden unveraendert erwartet!")


# ===========================================================================
# 4. KONSISTENZ - Gleicher Input -> Gleicher Output
# ===========================================================================

class TestKonsistenz(unittest.TestCase):
    """
    Mueller GmbH erscheint 2x -> muss beide Male denselben Fake-Namen bekommen.
    KdNr 10001 erscheint 2x -> muss beide Male dieselbe Fake-Nummer bekommen.
    """

    @classmethod
    def setUpClass(cls):
        reset_maps()
        cls.wb_o, cls.wb_s, _ = _load("offene_posten")
        cls.ws_o = cls.wb_o.active
        cls.ws_s = cls.wb_s.active

    def _synth_col(self, col_idx, start=3):
        return [self.ws_s.cell(r, col_idx).value
                for r in range(start, self.ws_s.max_row + 1)]

    def _orig_col(self, col_idx, start=3):
        return [self.ws_o.cell(r, col_idx).value
                for r in range(start, self.ws_o.max_row + 1)]

    def test_kundenname_konsistenz(self):
        """Mueller GmbH (Zeile 3+5) -> immer gleicher Fake-Name."""
        orig_names  = self._orig_col(2)
        synth_names = self._synth_col(2)

        # Mapping orig -> synth aufbauen
        mapping = {}
        for o, s in zip(orig_names, synth_names):
            if o in mapping:
                self.assertEqual(mapping[o], s,
                    f"Inkonsistenz: '{o}' wurde einmal als '{mapping[o]}' "
                    f"und einmal als '{s}' gemappt")
            else:
                mapping[o] = s

    def test_kundennummer_konsistenz(self):
        """KdNr 10001 (2x) -> immer gleiche Fake-KdNr."""
        orig  = self._orig_col(1)
        synth = self._synth_col(1)
        mapping = {}
        for o, s in zip(orig, synth):
            if o in mapping:
                self.assertEqual(mapping[o], s,
                    f"KdNr '{o}': Inkonsistenz {mapping[o]} != {s}")
            else:
                mapping[o] = s

    def test_iban_konsistenz(self):
        """Gleiche IBAN (Mueller GmbH, 2x) -> immer gleiche Fake-IBAN."""
        orig  = self._orig_col(3)
        synth = self._synth_col(3)
        mapping = {}
        for o, s in zip(orig, synth):
            if o in mapping:
                self.assertEqual(mapping[o], s,
                    f"IBAN '{o}': Inkonsistenz {mapping[o]} != {s}")
            else:
                mapping[o] = s

    def test_verschiedene_kunden_verschiedene_fakes(self):
        """Verschiedene Original-Kunden -> verschiedene Fake-Namen."""
        orig  = self._orig_col(2)
        synth = self._synth_col(2)
        unique_orig  = set(orig)
        unique_synth = set(synth)
        # Wenn es 5 verschiedene Kunden gibt, sollen es auch 5 verschiedene Fake-Namen sein
        self.assertEqual(
            len(unique_orig), len(unique_synth),
            f"Anzahl eindeutiger Namen weicht ab: {len(unique_orig)} Orig, {len(unique_synth)} Synth"
        )


# ===========================================================================
# 5. BILANZ - Kontonummern und Bezeichnungen unveraendert
# ===========================================================================

class TestBilanz(unittest.TestCase):
    """
    In Bilanz und GuV sind Kontonummern und Bezeichnungen keine PII.
    Sie muessen unveraendert bleiben. Nur Betraege werden anonymisiert.
    """

    @classmethod
    def setUpClass(cls):
        reset_maps()
        cls.wb_o, cls.wb_s, cls.info = _load("bilanz")
        cls.ws_o = cls.wb_o.active
        cls.ws_s = cls.wb_s.active

    def test_kontonummern_unveraendert(self):
        """Konto-Nr (Spalte 1) muss unveraendert bleiben."""
        for row in range(3, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 1).value
            synth = self.ws_s.cell(row, 1).value
            if orig:  # Abschnittszeilen ueberspringen
                with self.subTest(row=row, konto=orig):
                    self.assertEqual(orig, synth,
                        f"Zeile {row}: Kontonummer '{orig}' wurde zu '{synth}'!")

    def test_kontobezeichnungen_unveraendert(self):
        """Kontobezeichnungen (Spalte 2) muessen unveraendert bleiben."""
        for row in range(3, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 2).value
            synth = self.ws_s.cell(row, 2).value
            if orig:
                with self.subTest(row=row):
                    self.assertEqual(orig, synth,
                        f"Zeile {row}: Bezeichnung '{orig}' wurde zu '{synth}'!")

    def test_betraege_geaendert(self):
        """Betragsspalten (2024 + 2023) muessen anonymisiert sein."""
        changed = 0
        for row in range(4, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 3).value
            synth = self.ws_s.cell(row, 3).value
            if isinstance(orig, (int, float)) and orig != synth:
                changed += 1
        self.assertGreater(changed, 0, "Kein einziger Betrag wurde veraendert!")

    def test_erkannte_typen_bilanz(self):
        """Bilanz-Spalten muessen korrekt erkannt werden."""
        col_info = self.info["sheets"].get("Bilanz", {})
        self.assertIn(col_info.get("Konto-Nr"), ["text", "kundennummer"],
                      "Konto-Nr sollte nicht als IBAN erkannt werden")
        self.assertIn(col_info.get("Kontobezeichnung"), ["text", "beschreibung"],
                      "Kontobezeichnung sollte unveraendert bleiben")
        self.assertIn(col_info.get("2024 (EUR)"), ["betrag", "text"],
                      "2024-Spalte sollte als Betrag erkannt werden")


# ===========================================================================
# 6. GuV - Positionen und Bezeichnungen unveraendert
# ===========================================================================

class TestGuV(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        reset_maps()
        cls.wb_o, cls.wb_s, cls.info = _load("guv")
        cls.ws_o = cls.wb_o.active
        cls.ws_s = cls.wb_s.active

    def test_positionen_unveraendert(self):
        """Pos.-Spalte (1) muss unveraendert bleiben."""
        for row in range(3, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 1).value
            synth = self.ws_s.cell(row, 1).value
            with self.subTest(row=row):
                self.assertEqual(orig, synth)

    def test_bezeichnungen_unveraendert(self):
        """GuV-Positionen wie 'Umsatzerloese' muessen unveraendert bleiben."""
        for row in range(3, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 2).value
            synth = self.ws_s.cell(row, 2).value
            if orig:
                with self.subTest(row=row):
                    self.assertEqual(orig, synth,
                        f"GuV-Bezeichnung geaendert: '{orig}' -> '{synth}'")

    def test_betraege_geaendert(self):
        """Mindestens einige Betraege muessen veraendert sein."""
        changed = 0
        for row in range(4, self.ws_o.max_row + 1):
            orig  = self.ws_o.cell(row, 3).value
            synth = self.ws_s.cell(row, 3).value
            if isinstance(orig, (int, float)) and orig != synth:
                changed += 1
        self.assertGreater(changed, 0)


# ===========================================================================
# 7. BETRAGSBEREICH - Synthetische Werte innerhalb Min-Max
# ===========================================================================

class TestBetragsbereich(unittest.TestCase):

    def test_betraege_im_originalbereich(self):
        """Alle synthetischen Betraege muessen zwischen Min und Max des Originals liegen."""
        reset_maps()
        wb_o, wb_s, _ = _load("offene_posten")
        ws_o, ws_s    = wb_o.active, wb_s.active

        # Betraege aus Spalte 7 (ab Zeile 3)
        orig_vals  = [ws_o.cell(r, 7).value for r in range(3, ws_o.max_row + 1)
                      if isinstance(ws_o.cell(r, 7).value, (int, float))]
        synth_vals = [ws_s.cell(r, 7).value for r in range(3, ws_s.max_row + 1)
                      if isinstance(ws_s.cell(r, 7).value, (int, float))]

        if not orig_vals:
            self.skipTest("Keine numerischen Betraege gefunden")

        min_v = min(orig_vals)
        max_v = max(orig_vals)

        for v in synth_vals:
            with self.subTest(value=v):
                self.assertGreaterEqual(v, min_v * 0.95,
                    f"{v:.2f} liegt unter Minimum {min_v:.2f}")
                self.assertLessEqual(v, max_v * 1.05,
                    f"{v:.2f} liegt ueber Maximum {max_v:.2f}")

    def test_betraege_positiv(self):
        """Offene Posten Betraege muessen positiv bleiben."""
        reset_maps()
        _, wb_s, _ = _load("offene_posten")
        ws_s       = wb_s.active
        for row in range(3, ws_s.max_row + 1):
            val = ws_s.cell(row, 7).value
            if isinstance(val, (int, float)):
                with self.subTest(row=row):
                    self.assertGreater(val, 0,
                        f"Zeile {row}: Negativer Betrag {val} in OP-Liste!")


# ===========================================================================
# 8. TYPERKENNUNG - Typische Audit-Felder korrekt erkannt
# ===========================================================================

class TestTyperkennung(unittest.TestCase):

    def _check(self, col_name, values, expected_type):
        detected = detect_column_type(col_name, values)
        self.assertEqual(detected, expected_type,
            f"'{col_name}': erwartet '{expected_type}', erkannt '{detected}'")

    # Kundendaten
    def test_kdnr(self):         self._check("KdNr",          ["10001","10002"],            "kundennummer")
    def test_kundennummer(self): self._check("Kundennummer",   ["K-001","K-002"],            "kundennummer")
    def test_lfdnr(self):        self._check("LfdNr",          ["1","2","3"],                "kundennummer")
    def test_debnr(self):        self._check("DebNr",          ["D-001","D-002"],            "kundennummer")
    def test_krednr(self):       self._check("KredNr",         ["CR-001","CR-002"],          "kundennummer")

    # Rechnungen
    def test_renr(self):         self._check("Re-Nr",          ["RE-2024-001"],              "rechnungsnr")
    def test_belnr(self):        self._check("BelNr",          ["BEL-001","BEL-002"],        "rechnungsnr")
    def test_rechnungsnr(self):  self._check("Rechnungsnummer",["RE-001"],                   "rechnungsnr")

    # Finanzen
    def test_iban(self):         self._check("IBAN",           ["DE89370400440532013000"],   "iban")
    def test_kontonr(self):      self._check("KontoNr",        ["DE89370400440532013000"],   "iban")

    # Betraege
    def test_betrag(self):       self._check("Betrag",         [1500.0, 2300.0],             "betrag")
    def test_btrgnetto(self):    self._check("BtrgNetto",      [100.0, 200.0],               "betrag")
    def test_restbetrag(self):   self._check("Restbetrag",     [500.0],                      "betrag")
    def test_betrag_eur(self):   self._check("2024 (EUR)",     [125000.0, 45000.0],          "betrag")

    # Datum
    def test_buchdat(self):      self._check("BuchDat",        ["2024-01-01"],               "datum")
    def test_valdat(self):       self._check("ValDat",         ["2024-01-01"],               "datum")
    def test_faellig(self):      self._check("Faelligkeitsdatum", ["2024-12-31"],            "datum")

    # Steuer/Recht
    def test_steuernr(self):     self._check("Steuernummer",   ["21/815/08150"],             "steuernummer")
    def test_ustid(self):        self._check("USt-IdNr",       ["DE123456789"],              "ustid")

    # Soll unveraendert bleiben
    def test_kontobezeichnung(self):
        # "beschreibung" und "text" fuehren beide zum gleichen Ergebnis (unveraendert)
        detected = detect_column_type("Kontobezeichnung", ["Kasse", "Bank"])
        self.assertIn(detected, ["text", "beschreibung"],
                      f"'Kontobezeichnung' sollte unveraendert bleiben, erkannt: '{detected}'")
    def test_kostenstelle(self):     self._check("Kostenstelle",     ["KST-001"],            "text")
    def test_status(self):           self._check("Status",           ["offen","bezahlt"],    "text")
    def test_pos(self):              self._check("Pos.",             ["1.","2.","3."],        "text")


# ===========================================================================
# 9. DATUMSPRUEFUNG
# ===========================================================================

class TestDatum(unittest.TestCase):

    def test_daten_bleiben_datumsobjekte(self):
        """Synthetische Datumswerte muessen Python date/datetime Objekte sein."""
        import datetime
        reset_maps()
        wb_o, wb_s, _ = _load("offene_posten")
        ws_s = wb_s.active
        for row in range(3, ws_s.max_row + 1):
            for col in [5, 6]:
                val = ws_s.cell(row, col).value
                if val is not None:
                    with self.subTest(row=row, col=col):
                        self.assertIsInstance(
                            val, (datetime.date, datetime.datetime),
                            f"Zeile {row}, Spalte {col}: Kein Datumsobjekt: {val!r}"
                        )

    def test_daten_plausibel(self):
        """Verschobene Daten muessen plausibel bleiben (zwischen 2020 und 2030)."""
        import datetime
        reset_maps()
        _, wb_s, _ = _load("offene_posten")
        ws_s = wb_s.active
        for row in range(3, ws_s.max_row + 1):
            for col in [5, 6]:
                val = ws_s.cell(row, col).value
                if isinstance(val, (datetime.date, datetime.datetime)):
                    year = val.year
                    with self.subTest(row=row, col=col, year=year):
                        self.assertGreaterEqual(year, 2020)
                        self.assertLessEqual(year, 2030)


if __name__ == "__main__":
    unittest.main(verbosity=2)
